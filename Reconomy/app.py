from flask import Flask, render_template, request
from flask_mysqldb import MySQL
from waitress import serve
from openpyxl import Workbook, load_workbook, cell
import os.path
import datetime

wsgiapp = Flask(__name__, template_folder='templates', static_folder='static')

wsgiapp.config['MYSQL_HOST'] = 'localhost'
wsgiapp.config['MYSQL_USER'] = 'root'
wsgiapp.config['MYSQL_PASSWORD'] = ''
wsgiapp.config['MYSQL_DB'] = 'flaskapp'

mysql = MySQL(wsgiapp)

@wsgiapp.route('/')
def index():
    try:
        cur = mysql.connection.cursor()
        cur.execute('SELECT name FROM users')
        fetchdata = cur.fetchall()
        cur.close()
        username = request.args.get('todo')
        buttonStatus = request.args.get('button_status')
        textArea = request.args.get('messagecontent')
        currentTime = datetime.datetime.today()
        # strips spaces from username and check if it true
        if username and username.strip():
                # check status of button and call either checkout or checkin
                if buttonStatus == 'check-in':
                    xlsxCheckout(username, currentTime, textArea)
                elif buttonStatus == 'check-out':
                    xlsxCheckin(username, currentTime, textArea)
                else:
                    print('Error')
    except Exception as error:
        return repr(error)
    # NameList = readUser()
    return render_template('index.html', user = fetchdata)

# edit xlsx-file and checks out a user
def xlsxCheckout(user, time, comments):
    date = str(datetime.date.today())
    fileName = 'timeReport'+ date +'.xlsx'
    if os.path.isfile(fileName):
        try:
            Wb = load_workbook(fileName)
            Ws = Wb.active
            index = xlsxSearch(user)
            if index != False:
                index = 'C' + str(index)
                Ws[index].value = time
            else:
                Ws.append([user, time, None, comments])
            Wb.save(fileName)
        except Exception as Error:
            print(Error)
    else:
        print('Incorrect action')

# creates a xlsx-file and checks in a user
def xlsxCheckin(user, time, comments):
    date = str(datetime.date.today())
    fileName = 'timeReport'+ date +'.xlsx'
    if os.path.isfile(fileName):
        Wb = load_workbook(fileName)
        Ws = Wb.active
        Ws.append([user, time, None, comments])
        Wb.save(fileName)
    else:
        Wb = Workbook()
        Ws = Wb.active
        Ws.title = 'Data'
        Ws.append(['Name','Time checked in', 'Time checked out', 'Comment'])
        Ws.append([user, time, None, comments])
        Wb.save(fileName)

# requires to be called by check in/out function, Will search for a user and return index value of row
def xlsxSearch(user):
    date = str(datetime.date.today())
    fileName = 'timeReport' + date + '.xlsx'
    if os.path.isfile(fileName):
        try:
            Wb = load_workbook(fileName, read_only=True)
            Ws = Wb.active
            for rows in Ws.iter_rows(min_row=1, max_col=1):
                cellValue = rows[0].value
                if cellValue == user:
                    return rows[0].row
                else:
                    continue
            return False
        except Exception as error:
            print(error)
    else:
        print('File do not exist')
        return False

# used to run in debug mode
if __name__ == '__main__':
    serve(wsgiapp, host='127.0.0.1', port=80, url_scheme='https')
    # wsgiapp.run(debug=True)